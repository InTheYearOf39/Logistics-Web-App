from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from lmsapp.models import Package, User,DropPickZone, Warehouse, UserGoogleSheet
from django.contrib import messages
from lmsapp.forms import ChangePasswordForm, PackageForm, ExcelUploadForm
from django.contrib.auth import update_session_auth_hash
from django.core.mail import send_mail
from django.conf import settings
from django.views.decorators.clickjacking import xframe_options_exempt
import random
import string
from datetime import datetime, timedelta
from django.core.mail import send_mail, BadHeaderError
from django.db.models import Q
import pandas as pd
from django.http import HttpResponse
from lmsapp.utils import send_sms

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.db import IntegrityError
import gspread
from oauth2client.service_account import ServiceAccountCredentials



"""  
The view displays a warehouse dashboard template where warehouse users can see packages grouped by drop_pick_zone, 
select packages and assign them to available couriers
"""

def is_warehouse_user(user):
    return user.role == 'warehouse'

@login_required
@user_passes_test(is_warehouse_user)
def warehouse_dashboard(request):
    packages = Package.objects.filter(
        status='dropped_off',
        dropOffLocation__warehouse=request.user.warehouse
    ).select_related('dropOffLocation').order_by('dropOffLocation__tag')

    if request.method == 'POST':
        selected_packages = request.POST.getlist('selected_packages')
        courier_id = request.POST.get('selectCourier')  # Change the name attribute of the courier select field to "selectCourier"
        if selected_packages and courier_id:
            courier = get_object_or_404(User, id=courier_id, role='courier', status='available')
            packages = Package.objects.filter(id__in=selected_packages)
            
            for package in packages:
                if package.deliveryType == 'premium':
                    package.status = 'en_route'
                elif package.deliveryType == 'express':
                    package.status = 'in_transit'
                else:
                    package.status = 'dispatched'
                package.courier = courier
                package.save()

            courier_status = Package.objects.filter(courier=courier, status='dispatched').exists()

            if courier_status:
                courier.status = 'on-trip'
                courier.save()

            messages.success(request, 'Packages successfully assigned to courier.')
            return redirect('warehouse_dashboard')


    if request.method == 'GET':
        success_message = request.GET.get('success_message', "")
        # Filter available couriers based on the warehouse
        available_couriers = User.objects.filter(
            role='courier', status='available', warehouse=request.user.warehouse
        )

        context = {
            'packages': packages,
            'available_couriers': available_couriers,
            'success_message': success_message
        }

        return render(request, 'warehouse/warehouse_dashboard.html', context)


@login_required
@user_passes_test(is_warehouse_user)
def premium_dashboard(request):
    packages = Package.objects.filter(
        Q(deliveryType='premium'),
        status='upcoming',
        warehouse=request.user.warehouse
    ).select_related('dropOffLocation').order_by('dropOffLocation__tag')

    if request.method == 'POST':
        selected_packages = request.POST.getlist('selected_packages')
        courier_id = request.POST.get('selectCourier')
        if selected_packages and courier_id:
            courier = get_object_or_404(User, id=courier_id, role='courier', status='available')
            packages = Package.objects.filter(id__in=selected_packages)

            for package in packages:
                if package.deliveryType == 'premium':
                    package.status = 'en_route'
                elif package.deliveryType == 'express':
                    package.status = 'in_transit'
                else:
                    package.status = 'dispatched'
                package.courier = courier
                package.save()

            courier_status = Package.objects.filter(courier=courier, status='dispatched').exists()

            if courier_status:
                courier.status = 'on-trip'
                courier.save()

            messages.success(request, 'Packages successfully assigned to courier.')
            return redirect('warehouse_dashboard')

    # Filter couriers based on the warehouse
    available_couriers = User.objects.filter(
        role='courier', status='available', warehouse=request.user.warehouse
    )

    context = {
        'packages': packages,
        'available_couriers': available_couriers
    }

    return render(request, 'warehouse/premium_dashboard.html', context)


@login_required
@user_passes_test(is_warehouse_user)
def express_dashboard(request):
    packages = Package.objects.filter(
        Q(deliveryType='express'),
        status='upcoming',
        warehouse=request.user.warehouse
    ).select_related('dropOffLocation').order_by('dropOffLocation__tag')

    if request.method == 'POST':
        selected_packages = request.POST.getlist('selected_packages')
        courier_id = request.POST.get('selectCourier')  # Change the name attribute of the courier select field to "selectCourier"
        if selected_packages and courier_id:
            courier = get_object_or_404(User, id=courier_id, role='courier', status='available')
            packages = Package.objects.filter(id__in=selected_packages)
            
            for package in packages:
                if package.deliveryType == 'premium':
                    package.status = 'en_route'
                elif package.deliveryType == 'express':
                    package.status = 'in_transit'
                else:
                    package.status = 'dispatched'
                package.courier = courier
                package.save()

            courier_status = Package.objects.filter(courier=courier, status='dispatched').exists()

            if courier_status:
                courier.status = 'on-trip'
                courier.save()

            messages.success(request, 'Packages successfully assigned to courier.')
            return redirect('warehouse_dashboard')

    available_couriers = User.objects.filter(
        role='courier', status='available', warehouse=request.user.warehouse
    )

    context = {
        'packages': packages,
        'available_couriers': available_couriers
    }

    return render(request, 'warehouse/express_dashboard.html', context)



# def warehouse_dashboard(request):
#     # Assuming you have the warehouse information from the user or some other source
#     # Replace this with the correct warehouse ID
#     packages = Package.objects.all()

#     context = {
#         'packages': packages,
#     }

#     return render(request, 'warehouse/warehouse_dashboard.html', context)


"""
The view handles the change password functionality and renders a template with a form 
for users to enter their new password.
"""
@login_required
def change_password(request):
    if request.method == 'POST':
        form = ChangePasswordForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important to update the session
            messages.success(request, 'Your password has been successfully changed.')
            return redirect('warehouse_dashboard')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = ChangePasswordForm(request.user)
    return render(request, 'warehouse/change_password.html', {'form': form})

""" 
Handles the confirmation of package arrival at the warehouse. It updates the package status, 
updates the courier's status if applicable, sends an email  ication to the sender
"""
@login_required
@user_passes_test(is_warehouse_user)
def confirm_arrival(request, package_id):
    if request.method == 'POST':
        warehouse_user = request.user
        warehouse = warehouse_user.warehouse

        package = Package.objects.get(pk=package_id)
        package.status = 'in_house'
        courier = package.courier
        if courier:
            courier.status = 'available'
            courier.save()
        
        # Assign the warehouse to the package
        package.warehouse = warehouse
        
        # Save the package
        package.save()
        
        # Send email to sender
        subject = 'Package Dropped Off at Warehouse'
        message = f'Dear Customer, your package with delivery number {package.package_number} has been dropped off at the warehouse.'

        sender_user = User.objects.get(username=package.user.username)
        sender_email = sender_user.email

        send_mail(subject, message, settings.EMAIL_HOST_USER, [sender_email])   

        messages.success(request, "Package arrival notified successfully.")
    else:
        messages.error(request, "Invalid request.")

    return redirect('in_house')  # Replace with the appropriate URL for the warehouse dashboard

""" 
The view displays a list of packages with the 'in_house'status and allows the user to assign selected packages 
to available couriers and drop_pick_zones, and updates the package status to 'in_transit'.
""" 
@login_required
@user_passes_test(is_warehouse_user)
def in_house(request):
    ready_packages = Package.objects.filter(status='in_house')

    if request.method == 'POST':
        selected_packages = request.POST.getlist('selected_packages')
        courier_id = request.POST.get('courier')

        if selected_packages and courier_id:
            courier = get_object_or_404(User, id=courier_id, role='courier')

            # Update the packages with the assigned courier and change their status
            packages = Package.objects.filter(id__in=selected_packages)
            package_types = set(package.deliveryType for package in packages)
            status_updated = False

            for package in packages:
                if package.deliveryType == 'premium':
                    package.status = 'en_route'
                    status_updated = True
                elif package.deliveryType == 'express':
                    package.status = 'in_transit'
                    status_updated = True
                else:
                    package.status = 'in_transit'  # Change status to 'in_transit' for standard packages
                    status_updated = True
                package.courier = courier
                package.save()

            courier.status = 'on-trip'
            courier.save()

            if status_updated:
                if 'premium' in package_types and 'express' in package_types:
                    messages.success(request, 'Premium and Express packages successfully assigned to courier.')
                elif 'premium' in package_types:
                    messages.success(request, 'Premium packages successfully assigned to courier.')
                elif 'express' in package_types:
                    messages.success(request, 'Express packages successfully assigned to courier.')
                else:
                    messages.success(request, 'Standard packages successfully assigned to courier.')
            else:
                messages.error(request, 'No packages were assigned.')

            return redirect('in_house')
        else:
            messages.error(request, 'Please select packages and a courier.')

    context = {
        'ready_packages': ready_packages,
        'available_couriers': User.objects.filter(role='courier', status='available', warehouse=request.user.warehouse),
    }
    return render(request, 'warehouse/ready_packages.html', context)


""" 
The view displays a list of packages ready for pickup and allows the user to assign selected packages 
to couriers and drop_pick_zones. It handles form submissions, updates the package assignments and statuses, 
and provides available couriers and drop_pick_zones in the context.
""" 
@login_required
@user_passes_test(is_warehouse_user)
def ready_for_pickup(request):
    if request.method == 'POST':
        selected_packages = request.POST.getlist('selected_packages')
        courier_id = request.POST.get('courier')

        if selected_packages and courier_id:
            courier = get_object_or_404(User, id=courier_id, role='courier')

            # Update the packages with the assigned courier and change their status
            packages = Package.objects.filter(id__in=selected_packages, deliveryType=['premium', 'standard'])
            packages.update(courier=courier, status='in_transit')
            
            courier.status = 'on-trip'
            courier.save()

            messages.success(request, 'Premium packages successfully assigned to courier.')

            return redirect('in_house')

    ready_packages = Package.objects.filter(status='in_transit', deliveryType='premium')
    available_couriers = User.objects.filter(role='courier', status='available')

    context = {
        'ready_packages': ready_packages,
        'available_couriers': available_couriers
    }
    return render(request, 'warehouse/ready_for_pickup.html', context)


@login_required
@user_passes_test(is_warehouse_user)
def new_arrivals(request):
    arrived_packages = Package.objects.filter(status='warehouse_arrival')
    success_message = ""
    if request.method == 'POST':

        return redirect('ready_packages')
    elif request.method == 'GET':
        success_message = request.GET.get('success_message', "")
            
    context = {
        'arrived_packages': arrived_packages,
        "success_message": success_message,
    }
    
    # extract_google_sheet_data(request)
    return render(request, 'warehouse/new_arrivals.html', context)

def generate_package_number():
    prefix = 'PN'
    digits = ''.join(random.choices(string.digits, k=5))
    return f'{prefix}{digits}'


@login_required
@xframe_options_exempt 
@user_passes_test(is_warehouse_user)
def add_package(request):
    user_warehouse = None  
    
    senders = User.objects.filter(role='sender')
    if request.method == 'POST':
        form = PackageForm(request.POST)

        if form.is_valid():
            package = form.save(commit=False)

            # Automatically set the warehouse to the one the user belongs to
            user_warehouse = request.user.warehouse
           
            delivery_type = request.POST.get('deliveryType')
            recipient_latitude = request.POST.get('recipient_latitude')
            recipient_longitude = request.POST.get('recipient_longitude')
            
            package.sendersName = form.cleaned_data["sendersName"]
            package.sendersEmail = form.cleaned_data["sendersEmail"]
            package.sendersContact = form.cleaned_data["sendersContact"]
            package.packageName = form.cleaned_data["packageName"]
            package.packageDescription = form.cleaned_data["packageDescription"]
            package.deliveryType = delivery_type
            package.warehouse = user_warehouse
            package.recipientName = form.cleaned_data["recipientName"]
            package.recipientEmail = form.cleaned_data["recipientEmail"]
            package.recipientTelephone = form.cleaned_data["recipientTelephone"]
            package.recipientAddress = form.cleaned_data["recipientAddress"]
            package.recipientIdentification = form.cleaned_data["recipientIdentification"]
            package.genderType = form.cleaned_data["genderType"]
            package.recipientPickUpLocation = form.cleaned_data["recipientPickUpLocation"]
            package.user = request.user
            package.recipient_latitude = recipient_latitude
            package.recipient_longitude = recipient_longitude
            package.package_number = generate_package_number()
            package.created_by = request.user

            package.status = 'in_house'

            selected_user_id = request.POST.get('user')
            if selected_user_id:
                selected_user = User.objects.get(id=selected_user_id)

                package.sendersEmail = selected_user.email
                package.sendersName = selected_user.username
            package.save()


            subject = 'Package Registered'
            message = (
                f"Dear Customer, your package has been successfully registered.\n\n"
                f"Package Number: {package.package_number}\n"
                f"Recipient: {package.recipientName}\n"
                f"Recipient Address: {package.recipientAddress}\n"
                f"If you have any questions, please contact our customer support team.\n"
            )

            if package.warehouse:
                message += f"Drop-off Location: {package.warehouse.name}\n"

            message += f"Status: {package.status}"
            
            from_email = settings.DEFAULT_FROM_EMAIL
            recipient_list = [package.recipientEmail, package.sendersEmail]
                           
            sender_contact = str(package.sendersContact).strip()

            if len(sender_contact) == 10 and sender_contact.startswith('0'):
                sender_contact = '+256' + sender_contact[1:]

            send_sms([sender_contact], message, settings.AFRICASTALKING_SENDER)
           
            try:
                send_mail(subject, message, from_email, recipient_list, fail_silently=False)
            except BadHeaderError as e:
                # Handle the BadHeaderError
                print("Error: BadHeaderError:", str(e))
            except Exception as e:
                # Handle other exceptions
                print("Error:", str(e))

            return redirect('in_house')
        else:
            warehouse = Warehouse.objects.all()
            context = { 'form': form, 'user_warehouse': request.user.warehouse, 'warehouse': warehouse }
            return render(request, 'warehouse/add_package.html', context)
    else:
        form = PackageForm()
        user_warehouse = request.user.warehouse

    # Get the warehouse data to populate the warehouse dropdown
    warehouse = Warehouse.objects.all()
    context = {'form': form, 'warehouse': warehouse, 'user_warehouse': user_warehouse, 'senders': senders}
    return render(request, 'warehouse/add_package.html', context)

# def add_package(request):
#     msg = None
#     user_warehouse = None  # Initialize user_warehouse
    
#     senders = User.objects.filter(role='sender')
#     if request.method == 'POST':
#         form = PackageForm(request.POST)

#         if form.is_valid():
#             package = form.save(commit=False)
#             package.created_by = request.user
#             package.package_number = generate_package_number()

#             # Automatically set the warehouse to the one the user belongs to
#             user_warehouse = request.user.warehouse
#             if user_warehouse:
#                 package.warehouse = user_warehouse
#             else:
#                 # Handle the case where the user does not have a warehouse
#                 pass

#             recipient_latitude = request.POST.get('recipient_latitude')
#             recipient_longitude = request.POST.get('recipient_longitude')
#             package.recipient_latitude = recipient_latitude
#             package.recipient_longitude = recipient_longitude

#             package.status = 'in_house'

#             selected_user_id = request.POST.get('user')
#             if selected_user_id:
#                 selected_user = User.objects.get(id=selected_user_id)

#                 package.sendersEmail = selected_user.email
#                 package.sendersName = selected_user.username
#             package.save()

            
#             # Send an email to the sender
#             subject = 'Package Registered'
#             message = (
#                 f"Dear Customer, your package has been successfully registered.\n\n"
#                 f"Package Number: {package.package_number}\n"
#                 f"Recipient: {package.recipientName}\n"
#                 f"Recipient Address: {package.recipientAddress}\n"
#                 f"If you have any questions, please contact our customer support team.\n"
#             )

#             if package.dropOffLocation:
#                 message += f"Drop-off Location: {package.dropOffLocation.name}\n"

#             message += f"Status: {package.status}"
            
#             from_email = settings.DEFAULT_FROM_EMAIL
#             recipient_list = [package.recipientEmail, package.sendersEmail]
                           
#             sender_contact = str(package.sendersContact).strip()

#             if len(sender_contact) == 10 and sender_contact.startswith('0'):
#                 sender_contact = '+256' + sender_contact[1:]

#             send_sms([sender_contact], message, settings.AFRICASTALKING_SENDER)
           
#             try:
#                 send_mail(subject, message, from_email, recipient_list, fail_silently=False)
#             except BadHeaderError as e:
#                 # Handle the BadHeaderError
#                 print("Error: BadHeaderError:", str(e))
#             except Exception as e:
#                 # Handle other exceptions
#                 print("Error:", str(e))

#             msg = 'Package registered'
#             return redirect('in_house')
#         else:
#             msg = 'Form is not valid'
#     else:
#         form = PackageForm()
#         user_warehouse = request.user.warehouse

#     # Get the warehouse data to populate the warehouse dropdown
#     warehouse = Warehouse.objects.all()  # Adjust this based on your model
#     context = {'form': form, 'warehouse': warehouse, 'user_warehouse': user_warehouse, 'senders': senders}
#     return render(request, 'warehouse/add_package.html', context)

@login_required
@user_passes_test(is_warehouse_user)
def packages_delivered(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    req_action = request.GET.get('req_a',"")

    packages_delivered = Package.objects.filter(status='completed')
    
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Ensure that start_date is always before end_date
        if start_date > end_date:
            start_date, end_date = end_date, start_date
        
        # Adjust end_date to be inclusive of the entire day
        end_date += timedelta(days=1)
        
        packages_delivered = packages_delivered.filter(completed_at__range=(start_date, end_date))
    elif start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        packages_delivered = packages_delivered.filter(completed_at__date=start_date)
      
    # determine wether to export or render html
    if req_action == "excel":
         # Create a response with the Excel file
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Packages_delivered_summary_data.xlsx"'

        # Create a new workbook
        workbook = Workbook()
        # print(f"initital workbook: {workbook}")
        # Create a sheet for packages_delivered
        sheet_delivered = workbook.active
        sheet_delivered.title = 'Packages Delivered'
        columns_to_include_delivered = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]  # Include columns
        write_data_to_sheet(sheet_delivered, packages_delivered, columns_to_include_delivered)


        # Save the workbook to the response
        workbook.save(response)
        print(f"The work book: {workbook}")

        print(f"The response: {response}")
        return response
    else:
        context = {
            'packages_delivered': packages_delivered,
        }
        
        return render(request, 'warehouse/packages_delivered.html', context )

@login_required
@user_passes_test(is_warehouse_user)
def packages_received(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    req_action = request.GET.get('req_a',"")


    packages_received = Package.objects.filter(status__in=['in_house', 'ready_for_pickup'])
    
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Ensure that start_date is always before end_date
        if start_date > end_date:
            start_date, end_date = end_date, start_date
        
        # Adjust end_date to be inclusive of the entire day
        end_date += timedelta(days=1)
        
        packages_received = packages_received.filter(received_at__range=(start_date, end_date))
    elif start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        packages_received = packages_received.filter(received_at__date=start_date)
       
    # determine wether to export or render html
    if req_action == "excel":
        # Create a response with the Excel file
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Packages_received_summary_data.xlsx"'

        # Create a new workbook
        workbook = Workbook()

        # Create a sheet for packages_received
        sheet_received = workbook.active
        sheet_received.title = 'Packages Received'

        # Specify columns to include for packages_delivered_export view
        columns_to_include_received = [0, 1, 2, 3, 4, 5, 6, 7, 8, 10]  # Include columns
        write_data_to_sheet(sheet_received, packages_received, columns_to_include_received)

        # Save the workbook to the response
        workbook.save(response)

        return response
    
    else:        
    
        context = {
            'packages_received': packages_received,
        }
        
        return render(request, 'warehouse/packages_received.html', context )
        

def write_data_to_sheet(sheet, queryset, columns_to_include):
    # Write header row
    header = ['Package Name', 'Package Number', 'Sender Address', 'Sender Name', 'Recipient Name', 'Recipient Address', 'Package Description', 'Delivery Type', 'Status', 'Time Completed', 'Time Received']
    included_header = [header[i] for i in columns_to_include]
    sheet.append(included_header)

    # Write data rows
    for package in queryset:
        row_data = [
            package.packageName,
            package.package_number,
            package.sendersAddress,
            package.sendersName,
            package.recipientName,
            package.recipientAddress,
            package.packageDescription,
            package.deliveryType,
            package.status,
            package.completed_at,
            package.in_house_at
        ]
        included_row_data = [row_data[i] for i in columns_to_include]
        sheet.append(included_row_data)


def preprocess_phone_number(phone_number):
    phone_number_str = str(phone_number)
    cleaned_phone_number = ''.join(filter(str.isdigit, phone_number_str))
    # Add the leading zero if missing
    if not cleaned_phone_number.startswith('0'):
        cleaned_phone_number = '0' + cleaned_phone_number

    return cleaned_phone_number


@login_required
@user_passes_test(is_warehouse_user)
def upload_excel(request):
    username_list = ['Ivan', 'muhumuza', 'Shem', 'Izzy']

    senders = User.objects.filter(username__in=username_list, role='sender')

    if request.method == "POST":
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = form.cleaned_data['excel_file']
            wb = load_workbook(excel_file)
            sheet = wb.active

            # header_mapping = {
            #     "order_id": 1, 
            #     "order_date": 2, 
            #     "recipient_name": 3, 
            #     "delivery_address": 4, 
            #     "city": 5, 
            #     "phone": 6, 
            #     "item": 7
            # }

            header_mapping = {
                "order_id": "Order ID", 
                "order_date": "Order Date",  
                "recipient_name": "Name of\nReceiver", 
                "delivery_address": "Deliery address", 
                "city": "City", 
                "phone": "Phone", 
                "item": "Item1",
                "quantity": "QTY(pieces)"
            }

            for row in sheet.iter_rows(min_row=2, values_only=True):

                order_date_value = row[header_mapping["order_date"]]   
                if order_date_value is None:
                    # Handle the case where order_date_value is None (blank cell)
                    continue
                
                order_date_str = order_date_value.strftime("%Y-%m-%d")

                # Check if the phone field is blank
                recipient_telephone = preprocess_phone_number(row[header_mapping["phone"]])
                if not recipient_telephone:
                    recipient_telephone = "Not Provided"  

                recipient_address = row[header_mapping["delivery_address"]]
                city = row[header_mapping["city"]]

                if not recipient_address and not city:
                    recipient_address_city = ""
                elif not recipient_address:
                    recipient_address_city = city
                elif not city:
                    recipient_address_city = recipient_address
                else:
                    recipient_address_city = recipient_address + ', -' + city

                package_number = row[header_mapping["order_id"]]
                order_date = parse_date(order_date_str)

                # Check if a package with the same package_number already exists
                existing_package = Package.objects.filter(package_number=package_number).first()

                if existing_package:
                    # Package with the same package_number already exists, skip
                    continue
                
                
                user_id = request.POST.get('client')
                try:
                    user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    user = None

                # Check if user has the 'name' attribute, if not, use 'username'
                if user and hasattr(user, 'name') and user.name:
                    senders_name = user.name
                else:
                    senders_name = user.username

                package = Package(
                    user=user,
                    packageName=row[header_mapping["item"]],
                    deliveryType='premium',  
                    package_number=package_number,
                    recipientName=row[header_mapping["recipient_name"]],
                    recipientEmail='',  
                    recipientTelephone=recipient_telephone,
                    recipientAddress=recipient_address_city,
                    packageDescription='',
                    sendersName=senders_name,  
                    sendersEmail=user.email,  
                    sendersAddress=user.address,  
                    sendersContact='',  
                    created_on=order_date,
                    created_by=user,
                    modified_by=request.user,
                    assigned_at=timezone.now(),
                    status='warehouse_arrival',
                    warehouse=request.user.warehouse
                )

                try:
                    package.save()
                except IntegrityError as ie:
                    if "UNIQUE constraint" in str(ie):
                        # Handle integrity error when package_number is already taken
                        messages.error(request, "Error: The package number already exists.")
                    else:
                        messages.error(request, "An error occurred while saving the package.")

            return redirect("new_arrivals")
    else:
        form = ExcelUploadForm()

    return render(request, "warehouse/upload_excel.html", {"form": form, "senders": senders})



# @login_required
# @user_passes_test(is_warehouse_user)
# def extract_google_sheet_data(request):
#     # username_list = ['Ivan', 'muhumuza', 'Shem', 'Izzy']

#     # senders = User.objects.filter(username__in=username_list, role='sender')

#     scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive',
#             'https://www.googleapis.com/auth/drive.file', 'https://www.googleapis.com/auth/spreadsheets']


#     credentials = ServiceAccountCredentials.from_json_keyfile_name('cred.json', scope)

#     client = gspread.authorize(credentials)

#     sheet = client.open_by_url("https://docs.google.com/spreadsheets/d/1xsk1k8F1rJJ4nj350LGSJpRXUIBv_F-t6KPIyGdCJ_I/edit#gid=0").sheet1

#     google_sheets_data = sheet.get_all_records()

#     header_mapping = {
#         "order_id": "Order ID", 
#         "order_date": "Order Date",  
#         "recipient_name": "Name of\nReceiver", 
#         "delivery_address": "Deliery address", 
#         "city": "City", 
#         "phone": "Phone", 
#         "item": "Item1",
#         "quantity": "QTY(pieces)"
#     }

#     for row in google_sheets_data:

#         order_date_value = row[header_mapping["order_date"]]
#         print("Order Date Value:", order_date_value) 

#         if order_date_value:
#             order_date = datetime.strptime(order_date_value, "%d/%b/%Y")
#             order_date_str = order_date.strftime("%Y-%m-%d")
#         else:
#             # Handle the case where order_date_value is empty
#             # You can choose to skip processing this row or handle it according to your logic
#             continue
        

#         # Assuming order_date_value is a string in the format "YYYY-MM-DD"
#         order_date_str = order_date_value
#         order_date = datetime.strptime(order_date_str, "%d/%b/%Y")

#         # Check if the phone field is blank
#         recipient_telephone = preprocess_phone_number(row[header_mapping["phone"]])
#         if not recipient_telephone:
#             recipient_telephone = "Not Provided"  

#         recipient_address = row[header_mapping["delivery_address"]]
#         city = row[header_mapping["city"]]

#         if not recipient_address and not city:
#             recipient_address_city = ""
#         elif not recipient_address:
#             recipient_address_city = city
#         elif not city:
#             recipient_address_city = recipient_address
#         else:
#             recipient_address_city = recipient_address + ' --' + city

#         package_number = row[header_mapping["order_id"]]

#         # Check if a package with the same package_number already exists
#         existing_package = Package.objects.filter(package_number=package_number).first()

#         if existing_package:
#             # Package with the same package_number already exists, skip
#             continue
        
#         # Get the item name and quantity from the row
#         item_name = row[header_mapping["item"]]
#         quantity = row[header_mapping["quantity"]]

#         # Append the quantity to the item name if quantity is greater than 1
#         if quantity > 1:
#             item_name += f' *({quantity} units)'

#         # user_id = request.POST.get('client')
#         user_id = 26
#         try:
#             user = User.objects.get(id=user_id)
#         except User.DoesNotExist:
#             user = None

#         # Check if user has the 'name' attribute, if not, use 'username'
#         if user and hasattr(user, 'name') and user.name:
#             senders_name = user.name
#         else:
#             senders_name = user.username

#         package = Package(
#             user=user,
#             packageName=item_name,
#             deliveryType='premium',  
#             package_number=package_number,
#             recipientName=row[header_mapping["recipient_name"]],
#             recipientEmail='',  
#             recipientTelephone=recipient_telephone,
#             recipientAddress=recipient_address_city,
#             packageDescription='',
#             sendersName=senders_name,  
#             sendersEmail=user.email,  
#             sendersAddress=user.address,  
#             sendersContact='',  
#             created_on=order_date,
#             created_by=user,
#             modified_by=request.user,
#             assigned_at=timezone.now(),
#             status='warehouse_arrival',
#             warehouse=request.user.warehouse
#         )

#         try:
#             package.save()
#         except IntegrityError as ie:
#             if "UNIQUE constraint" in str(ie):
#                 # Handle integrity error when package_number is already taken
#                 messages.error(request, "Error: The package number already exists.")
#             else:
#                 messages.error(request, "An error occurred while saving the package.")

#     return redirect("new_arrivals")


def is_empty_value(val):
    if val is None:
        return True
    elif str(val).strip().__len__() == 0:
        return True
    else:
        return False

@login_required
# @user_passes_test(is_warehouse_user)
def extract_google_sheet_data(request):

    sheets = UserGoogleSheet.objects.all()

    if request.method == "POST":
        
        sheet_id = request.POST.get('sheet_id')
        empty_rows = 0
        data_rows = 0
        skipped_rows = 0
        inserted_rows = 0
        user_google_sheet = None
        user = None

        try:
            user_google_sheet = UserGoogleSheet.objects.get(id=sheet_id)
            user = user_google_sheet.user
        except UserGoogleSheet.DoesNotExist:
            return HttpResponse("No Google Sheet URL found for this user.")

        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive',
                'https://www.googleapis.com/auth/drive.file', 'https://www.googleapis.com/auth/spreadsheets']

        credentials = ServiceAccountCredentials.from_json_keyfile_name('cred.json', scope)
        client = gspread.authorize(credentials)

        sheet = client.open_by_url(user_google_sheet.google_sheet_url).sheet1

        google_sheets_data = sheet.get_all_records()

        all_fields = ["created_on", "package_number", "packageName", "sendersName"
                      , "packageDescription", "sendersName", "sendersEmail", "sendersAddress", "sendersContact"
                      , "sender_latitude", "sender_longitude"
                      , "recipientName", "recipientEmail", "recipientTelephone"
                      , "recipientAddress", "recipientIdentification", "recipient_latitude"
                      , "recipient_longitude", "genderType"
                      , "deliveryFee", "deliveryType"]
        
        
        mapping_options = user_google_sheet.header_mapping
        header_mapping = mapping_options["fields"]
        mapping_settings = mapping_options["settings"]
        mapping_defaults = mapping_options.get("defaults", {})
        custom_fields = mapping_options.get("custom_fields", [])

        # a row is considered not empty if it has a value in any of these
        non_empty_indicator_field = mapping_settings["non_empty_indicator_field"]
        # thse will be validated for each non empty row
        all_fields_mandatory = mapping_settings["all_fields_mandatory"]
        
        

        try:
            # save batch insert data
            batch_list = list()
            # get a list of all package number
            # advised to save original pacakge number
            existing_package_numbers = [ a_pkg.package_number for a_pkg in Package.objects.all() ]
            created_on_formats = mapping_settings["created_on_formats"]
            row_ind = 2 # offset heading row
            for row in google_sheets_data:
                # get all values be nullbale for those missing
                package_row = {}
                # assume its empty
                is_empty = True
                # loop through all possible pacakge fields
                # read all other fields
                for a_fld in non_empty_indicator_field:
                    # check if it is provided for in mappings
                    if header_mapping.get(a_fld, None) is not None:
                        # try to read it
                        package_row[a_fld] = row.get(header_mapping[a_fld], None)
                    else:
                        # if it was not in mapping, set to respective empty values
                        # definbe case for those to be empty bstrings
                        package_row[a_fld] = None
                    # is any of these fields has a value mark as not empty
                    if is_empty_value(package_row[a_fld]) is False:
                        is_empty = False

                # print("empty:"+ str(is_empty), batch_list.__len__())
                # if row not empty continue
                if is_empty is False:
                    data_rows += 1

                    # loop through all possible pacakge fields
                    # read all other fields
                    for a_fld in all_fields:
                        # check if it is provided for in mappings
                        if header_mapping.get(a_fld, None) is not None:
                            # try to read it
                            package_row[a_fld] = row.get(header_mapping[a_fld], None)
                        else:
                            # if it was not in mapping, set to respective empty values
                            # definbe case for those to be empty bstrings
                            # apply default value if available
                            package_row[a_fld] = mapping_defaults.get(a_fld, None)

                    # loop through all possible any custom fields
                    # read if any more cutsom fields
                    if custom_fields.__len__() > 0:
                        for a_fld in custom_fields:
                            # check if it is provided for in mappings
                            if header_mapping.get(a_fld, None) is not None:
                                # try to read it
                                package_row[a_fld] = row.get(header_mapping[a_fld], None)
                            else:
                                # if it was not in mapping, set to respective empty values
                                # definbe case for those to be empty bstrings
                                # apply default value if available
                                package_row[a_fld] = mapping_defaults.get(a_fld, None)
                        
                    
                    
                    # validate mandatory
                    for a_fld in all_fields_mandatory:
                        if is_empty_value(package_row[a_fld]) is True:
                            status_msg = "Value for {} on row {} should not be empty".format(header_mapping[a_fld], row_ind)
                            # print("mandatory missing")
                            return render(request, "warehouse/extract_google_sheet_data.html", {"sheets": sheets
                                                                                , "status_msg": status_msg  })

                    #### handle other data conversions
                    

                    #### handle date conversions if not date format yet
                    if is_empty_value(package_row["created_on"]) is False:
                        # only do this if not a valid datetime object yet
                        if not isinstance(package_row["created_on"], datetime):
                            for date_format in created_on_formats:
                                try:
                                    package_row["created_on"] = datetime.strptime(package_row["created_on"], date_format)
                                    # stop trying if it succeds for first format
                                    break
                                except Exception as ie:
                                    status_msg = "exception:{}".format(row_ind,str(ie) )
                            # check if up to now no valid conversion happed
                            if not isinstance(package_row["created_on"], datetime):
                                status_msg = "Could not convert date value {} on row {}: {}".format(package_row["created_on"], row_ind,str(ie) )
                                # print("date conversion failed")
                                return render(request, "warehouse/extract_google_sheet_data.html", {"sheets": sheets})
                            
                    ####default package date to now unless provided
                    if is_empty_value(package_row["created_on"]) is True:
                        package_row["created_on"] = datetime.now()

                    ##### clean phone number
                    if not package_row["recipientTelephone"]:
                        package_row["recipientTelephone"] = ""
                    package_row["recipientTelephone"] = preprocess_phone_number(package_row["recipientTelephone"])
                      

                    # add special case for city UNTILL WE ADD IT INTO THE DATBASE
                    ##### Add city to address info ########
                    # as long as city data exists for this row
                    if package_row.get("city", None) is not None:
                            package_row["recipientAddress"] = str(package_row["recipientAddress"]) + ' --' + str(package_row["city"])

                    ##### Add city to address info ########

                    

                    # add special case for quantity UNTILL WE ADD IT INTO THE DATBASE
                    ##### Add quantity to package name ########
                    # as long as quantity data exists for this row
                    if package_row.get("quantity", None) is not None:
                            package_row["packageName"] = str(package_row["packageName"]) + ' (' + str(package_row["quantity"]) + ")"

                    ##### Add quantity to package name ########


                    ###Set Sender name if not porvided
                    if is_empty_value(package_row["sendersName"]) is True:
                        # Check if user has the 'name' attribute, if not, use 'username'
                        if user and hasattr(user, 'name') and user.name:
                            package_row["sendersName"] = user.name
                        else:
                            package_row["sendersName"] = user.username
                    if is_empty_value(package_row["sendersEmail"]) is True:
                        package_row["sendersEmail"] = user.email
                    if is_empty_value(package_row["sendersAddress"]) is True:
                        package_row["sendersAddress"] = user.address

                    # print(package_row["package_number"])
                    #### check unique package number
                    ### NEeed to later csale it to preserver original pcakag enumber and add filed for extrernal_source_ids
                    if package_row["package_number"] in existing_package_numbers:
                        skipped_rows += 1
                    else:
                        inserted_rows += 1
                        # add to batch to insert later
                        batch_list.append(
                            Package(
                            user=user,
                            packageName=package_row["packageName"],
                            deliveryType=package_row["deliveryType"],  
                            package_number=package_row["package_number"],
                            recipientName=package_row["recipientName"],
                            recipientEmail=package_row["recipientEmail"],  
                            recipientTelephone=package_row["recipientTelephone"],
                            recipientAddress=package_row["recipientAddress"],
                            packageDescription=package_row["packageDescription"],
                            sendersName=package_row["sendersName"],  
                            sendersEmail=package_row["sendersEmail"],  
                            sendersAddress=package_row["sendersAddress"],  
                            sendersContact=package_row["sendersContact"],  
                            created_on=package_row["created_on"],
                            created_by=user,
                            modified_by=request.user,
                            assigned_at=timezone.now(),
                            status='warehouse_arrival',
                            warehouse=request.user.warehouse
                        )
                        )

                    # go next
                    row_ind += 1
                else:
                    # if empty row
                    # go next
                    row_ind += 1
                    empty_rows += 1
            
            # do the insert
            if batch_list.__len__() > 0:
                # import all at once but in btatches of 1000 each time
                Package.objects.bulk_create(batch_list, 1000)
                #for a_pkg in batch_list:
                #    a_pkg.save()
        except Exception as e:
            status_msg = "Failed to complete operation: {}".format(str(e))
            return render(request, "warehouse/extract_google_sheet_data.html", {"sheets": sheets
                                                                              , "status_msg": status_msg  })
        success_message = f"inserted:{inserted_rows}, skipped:{skipped_rows}" 
        # print(empty_rows, data_rows, skipped_rows, inserted_rows)
        return redirect(reverse('new_arrivals') + f'?success_message={success_message}')
    else:
        
        return render(request, "warehouse/extract_google_sheet_data.html", {"sheets": sheets})
    
@login_required
def courier_tracking(request):
    couriers = User.objects.filter(role='courier', warehouse=request.user.warehouse)
    return render(request, 'warehouse/courier_tracking.html', {'couriers': couriers})
